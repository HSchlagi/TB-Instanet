import openpyxl
from datetime import datetime, timedelta
import random

def create_demo_excel():
    """Erstellt eine Demo-Excel-Datei mit mehreren Zählpunkten"""
    
    # Neue Arbeitsmappe erstellen
    wb = openpyxl.Workbook()
    
    # Zeitreihe für einen Tag (15-Minuten-Intervalle)
    start_time = datetime(2024, 1, 1, 0, 0, 0)
    time_series = []
    for i in range(96):  # 24 Stunden * 4 (15-Minuten-Intervalle)
        time_series.append(start_time + timedelta(minutes=15*i))
    
    # Zählpunkt 1: Hauptlast (Gebäude)
    main_load = []
    for i, time in enumerate(time_series):
        # Tageslast-Profil: Morgens und abends höher
        hour = time.hour
        if 6 <= hour <= 8:  # Morgen
            base_load = 25 + random.uniform(-2, 2)
        elif 18 <= hour <= 22:  # Abend
            base_load = 30 + random.uniform(-3, 3)
        elif 23 <= hour or hour <= 5:  # Nacht
            base_load = 5 + random.uniform(-1, 1)
        else:  # Tag
            base_load = 15 + random.uniform(-2, 2)
        
        main_load.append(max(0, base_load))
    
    # Zählpunkt 2: Produktion (Maschinen)
    production_load = []
    for i, time in enumerate(time_series):
        hour = time.hour
        if 7 <= hour <= 17:  # Arbeitszeit
            base_load = 20 + random.uniform(-5, 5)
        else:  # Nachtschicht minimal
            base_load = 2 + random.uniform(-1, 1)
        
        production_load.append(max(0, base_load))
    
    # Zählpunkt 3: Klimaanlage
    hvac_load = []
    for i, time in enumerate(time_series):
        hour = time.hour
        if 8 <= hour <= 18:  # Bürozeiten
            base_load = 15 + random.uniform(-2, 2)
        else:  # Nachts aus
            base_load = 1 + random.uniform(-0.5, 0.5)
        
        hvac_load.append(max(0, base_load))
    
    # Zählpunkt 4: Beleuchtung
    lighting_load = []
    for i, time in enumerate(time_series):
        hour = time.hour
        if 6 <= hour <= 22:  # Beleuchtung an
            base_load = 8 + random.uniform(-1, 1)
        else:  # Nachts aus
            base_load = 1 + random.uniform(-0.3, 0.3)
        
        lighting_load.append(max(0, base_load))
    
    # Hauptlast-Blatt
    ws1 = wb.active
    ws1.title = "Hauptlast"
    ws1['A1'] = "Zeitstempel"
    ws1['B1'] = "Hauptlast_kW"
    
    for i, (time, load) in enumerate(zip(time_series, main_load), 2):
        ws1[f'A{i}'] = time.strftime("%Y-%m-%d %H:%M:%S")
        ws1[f'B{i}'] = round(load, 2)
    
    # Produktion-Blatt
    ws2 = wb.create_sheet("Produktion")
    ws2['A1'] = "Zeitstempel"
    ws2['B1'] = "Produktion_kW"
    
    for i, (time, load) in enumerate(zip(time_series, production_load), 2):
        ws2[f'A{i}'] = time.strftime("%Y-%m-%d %H:%M:%S")
        ws2[f'B{i}'] = round(load, 2)
    
    # Klimaanlage-Blatt
    ws3 = wb.create_sheet("Klimaanlage")
    ws3['A1'] = "Zeitstempel"
    ws3['B1'] = "Klimaanlage_kW"
    
    for i, (time, load) in enumerate(zip(time_series, hvac_load), 2):
        ws3[f'A{i}'] = time.strftime("%Y-%m-%d %H:%M:%S")
        ws3[f'B{i}'] = round(load, 2)
    
    # Beleuchtung-Blatt
    ws4 = wb.create_sheet("Beleuchtung")
    ws4['A1'] = "Zeitstempel"
    ws4['B1'] = "Beleuchtung_kW"
    
    for i, (time, load) in enumerate(zip(time_series, lighting_load), 2):
        ws4[f'A{i}'] = time.strftime("%Y-%m-%d %H:%M:%S")
        ws4[f'B{i}'] = round(load, 2)
    
    # Datei speichern
    wb.save('demo_load_profile.xlsx')
    
    print("✅ Demo-Excel-Datei erstellt: demo_load_profile.xlsx")
    print("📊 Zählpunkte:")
    print("  - Hauptlast: Gebäudelast (25-30 kW Peak)")
    print("  - Produktion: Maschinen (20 kW Arbeitszeit)")
    print("  - Klimaanlage: HVAC-System (15 kW Bürozeiten)")
    print("  - Beleuchtung: Licht (8 kW Tag)")
    print("")
    print("💡 Verwende diese Datei zum Testen des Excel-Imports!")

if __name__ == "__main__":
    create_demo_excel() 